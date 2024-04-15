# -*- coding: utf-8 -*-
"""
Created on Thu May 11 19:46:44 2023

@author: USER
"""

# 抓取指定欄位內容

import openpyxl

file = 'school.xlsx'

wb = openpyxl.load_workbook(file)

ws = wb['students']

for i in range(1, ws.max_column+1): # +1 否則最後一個抓不到
    print(ws.cell(column=i, row=5).value, end=' ')
    
print()

# 抓取區域欄位內容

for i in range(2, ws.max_row+1):
    for j in range(1, ws.max_column+1):
        print(ws.cell(row = i, column = j).value, end=' ')
        #print(ws.cell(i,j).value, end=' ')
    print()
    
    
 # 寫法二: 將整列抓出後，再一個一個給
for row in ws.rows:
    for cell in row:
        print(cell.value, end=' ')
    print()
        