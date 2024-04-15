# -*- coding: utf-8 -*-
"""
Created on Thu May 11 19:10:54 2023

@author: USER
"""

# pip install openpyxl -> spyder 不用另外裝，其他程式則要先裝此 moudle，非預設

import openpyxl # 不須再用 with open 開啟檔案，直接使用此moudle即可

file = 'school.xlsx'

wb = openpyxl.load_workbook(file)

# 抓取預設工作表名稱
ws = wb.active
print('目前工作表:', ws.title)

# 抓取所選工作表名稱(建議使用，因為可直接抓取所需要的內容)
ws = wb['students']
print('所選工作表名稱:', ws.title)

# 印出指定儲存格的值
print('A2 儲存格值:', ws['A2'].value)
print('B2 儲存格值:', ws['B2'].value)
print('C2 儲存格值:', ws['C2'].value)

# 印出指定儲存格在第幾欄第幾列
print(ws['A2'].column, ws['A2'].row)
print(ws['C1'].column, ws['C1'].row)

# 印出總欄數、總列數
print('工作表欄數:', ws.max_column)
print('工作表列數:', ws.max_row)

