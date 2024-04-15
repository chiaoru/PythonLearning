# -*- coding: utf-8 -*-
"""
Created on Thu May 11 20:54:36 2023

@author: USER
"""

import openpyxl

# 建立空白迴頁簿
wb = openpyxl.Workbook()

ws = wb.active

# 寫法一
ws['A1'] = '姓名'
ws['A2'] = 'Bill'
ws['B1'] = '學校'
ws['B2'] = '聯成'

# 寫法二:串列
rowdata = ['David', '中科大']
ws.append(rowdata)

# 寫法三
data = [['John', '中興'],['Mary','靜宜'],['Sue', '東海']]

for item in data:
    ws.append(item)
    


wb.save('lcc2.xlsx')