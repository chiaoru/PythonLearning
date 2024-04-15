# -*- coding: utf-8 -*-
"""
Created on Thu May 11 20:54:10 2023

@author: USER
"""

import openpyxl

wb = openpyxl.Workbook()

ws = wb.active

ws.title = 'member'


ws['A1'] = "姓名"
ws['B1'] = "學校"

rowdata = ['David','中科大']
ws.append(rowdata)

rowdata = ['Geogeo','中興']
ws.append(rowdata)

wb.save('lcc2.xlsx')