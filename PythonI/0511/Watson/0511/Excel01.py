# -*- coding: utf-8 -*-
"""
Created on Thu May 11 19:06:33 2023

@author: USER
"""

import openpyxl

file = 'school.xlsx'

wb = openpyxl.load_workbook(file)

# 抓取預設工作表名稱
ws = wb.active

print('工作表名稱：',ws.title)

ws = wb['students']

print('所選工作表名稱：',ws.title)

print('A2的值：',ws['A2'].value)
print('B2的值：',ws['B2'].value)
print('C的值：',ws['C2'].value)

print()
print(ws['A2'].column,ws['A2'].row)
print(ws['C1'].column,ws['C1'].row)

print('總欄數：',ws.max_column)
print('總列數：',ws.max_row)






