# -*- coding: utf-8 -*-
"""
Created on Thu May 11 19:46:05 2023

@author: USER
"""


import openpyxl

file = 'school.xlsx'

wb = openpyxl.load_workbook(file)

ws = wb['students']

for i in range(1,ws.max_column + 1):
    print(ws.cell(column=i,row=5).value,end=' ')
    
print()    


for i in range(2,ws.max_row + 1):
    for j in range(1,ws.max_column + 1):
        print(ws.cell(i,j).value , end=' ')
        
    print()    
        
print("第二種寫法")        

for row in ws.rows:

    for cell in row:
        print(cell.value,end=' ')
    print()    
        



